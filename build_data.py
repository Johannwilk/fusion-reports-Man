import openpyxl, json
from collections import defaultdict

SRC = '/mnt/user-data/outputs/Master_Flat_File_CLEAN.xlsx'
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['Tidy Data']
headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
idx = {h: i for i, h in enumerate(headers)}

rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if r[0] is None:
        continue
    rows.append(r)

print('total rows', len(rows))

# compute week-in-month rank per (fiscal_year, fiscal_month) based on ascending distinct ISO week
fy_fm_weeks = defaultdict(set)
for r in rows:
    fy = r[idx['Fiscal Year']]
    fm = r[idx['Fiscal Month']]
    iso = r[idx['ISO Week']]
    fy_fm_weeks[(fy, fm)].add(iso)

week_rank = {}
for (fy, fm), weeks in fy_fm_weeks.items():
    for rank, w in enumerate(sorted(weeks), start=1):
        week_rank[(fy, fm, w)] = rank

max_rank = max(week_rank.values())
print('max week rank found:', max_rank)

def region_bucket(region):
    return 'Durban' if region == 'DURBAN' else 'Atlantis'

out_rows = []
for r in rows:
    region = r[idx['Region']]
    category = r[idx['Category']]
    fy = r[idx['Fiscal Year']]
    fm = r[idx['Fiscal Month']]
    iso = r[idx['ISO Week']]
    wk = week_rank[(fy, fm, iso)]
    cust = r[idx['Customer Name']]
    qty = r[idx['Qty']] or 0
    total = r[idx['Total Selling']] or 0
    out_rows.append([
        region_bucket(region), category, fy, fm, wk, cust, round(qty,2), round(total,2)
    ])

print('sample', out_rows[0])
print('encoded rows', len(out_rows))

with open('/home/claude/master_rows.json', 'w') as f:
    json.dump(out_rows, f)

# distinct categories & fiscal years/months present for filter UI
cats = sorted(set(r[1] for r in out_rows))
fys = sorted(set(r[2] for r in out_rows))
print('categories', cats)
print('fiscal years', fys)

# fiscal month -> calendar month name mapping (FY starts March)
fm_names = {1:'March',2:'April',3:'May',4:'June',5:'July',6:'August',7:'September',8:'October',9:'November',10:'December',11:'January',12:'February'}
print(fm_names)
